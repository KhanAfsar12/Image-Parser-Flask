import os
import pandas as pd
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from openpyxl.drawing.image import Image as OpenpyxlImage
from flask import Flask, request, render_template, flash, redirect, url_for

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///questions.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/upload/'
app.config['SECRET_KEY'] = 'afsarkhan'

db = SQLAlchemy(app)

class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    previous_year = db.Column(db.String(50))
    level = db.Column(db.String(50))
    question_text = db.Column(db.String(255))
    question_image = db.Column(db.String(255), nullable=True)
    option1_text = db.Column(db.String(255), nullable=True)
    option1_image = db.Column(db.String(255), nullable=True)
    option2_text = db.Column(db.String(255), nullable=True)
    option2_image = db.Column(db.String(255), nullable=True)
    option3_text = db.Column(db.String(255), nullable=True)
    option3_image = db.Column(db.String(255), nullable=True)
    option4_text = db.Column(db.String(255), nullable=True)
    option4_image = db.Column(db.String(255), nullable=True)
    explanation = db.Column(db.Text)
    answer = db.Column(db.String(50))

with app.app_context():
    db.create_all()

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            import_from_excel(file_path)

            flash('Data has been imported successfully', 'success')
            return redirect(url_for('upload_file'))

    return render_template('upload.html')

def save_image_from_excel(image, folder, image_name):
    if image:
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        
        image_extension = image.path.split('.')[-1] 
        image_path = os.path.join(folder_path, f"{image_name}.{image_extension}")
        
        with open(image_path, 'wb') as img_file:
            img_file.write(image._data()) 

        return f'uploads/{folder}/{image_name}.{image_extension}'
    return None

    

def import_from_excel(file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb.active

    images_by_row = {image.anchor._from.row + 1: image for image in sheet._images}

    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False), start=2):
        previous_year = row[0].value
        level = row[1].value

        question_text = row[2].value if isinstance(row[2].value, str) else None
        option1_text = row[3].value if isinstance(row[3].value, str) else None
        option2_text = row[4].value if isinstance(row[4].value, str) else None
        option3_text = row[5].value if isinstance(row[5].value, str) else None
        option4_text = row[6].value if isinstance(row[6].value, str) else None
        explanation = row[7].value
        answer = row[8].value

        question_image_path = None
        option1_image_path = None
        option2_image_path = None
        option3_image_path = None
        option4_image_path = None

        if i in images_by_row:
            image = images_by_row[i]
            col_idx = image.anchor._from.col

            if col_idx == 2: 
                question_image_path = save_image_from_excel(image, 'questions', f'question_{i}')
            elif col_idx == 3:  
                option1_image_path = save_image_from_excel(image, 'options', f'option1_{i}')
            elif col_idx == 4: 
                option2_image_path = save_image_from_excel(image, 'options', f'option2_{i}')
            elif col_idx == 5:  
                option3_image_path = save_image_from_excel(image, 'options', f'option3_{i}')
            elif col_idx == 6:  
                option4_image_path = save_image_from_excel(image, 'options', f'option4_{i}')

        question = Question(
            previous_year=previous_year,
            level=level,
            question_text=question_text,
            question_image=question_image_path,
            option1_text=option1_text,
            option1_image=option1_image_path,
            option2_text=option2_text,
            option2_image=option2_image_path,
            option3_text=option3_text,
            option3_image=option3_image_path,
            option4_text=option4_text,
            option4_image=option4_image_path,
            explanation=explanation,
            answer=answer
        )
        db.session.add(question)

    db.session.commit()





@app.route('/questions')
def display():
    questions = Question.query.all()
    return render_template('questions.html', questions=questions)

if __name__ == "__main__":
    app.run(debug=True)
